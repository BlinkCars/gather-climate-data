# encoding:utf-8
from openpyxl import Workbook
import os
import math

stationArray = [56038, 56146, 56167, 56251, 56257]
stationDict = {
    '56038': '石渠县',
    '56146': '甘孜县',
    '56167': '道孚县',
    '56251': '新龙县',
    '56257': '理塘县',
}
factorArray = ['PRS', 'TEM', 'RHU', 'PRE', 'EVP', 'WIN', 'SSD', 'GST']
factorDict = {
    'PRS': ['气压', 4, 10, 0],
    'TEM': ['气温', 4, 10, 0],
    'RHU': ['相对湿度', 4, 9, 0],
    'PRE': ['降水', 4, 10, 0],
    'EVP': ['蒸发', 4, 9, 0],
    'WIN': ['风向风速', 4, 12, 0],
    'SSD': ['日照', 4, 8, 0],
    'GST': ['0cm 地温', 4, 10, 0],
}
# 结束时间
endDate = [2015, 12, 31]
# 线性相关所用数组
arr_evp_x = []
arr_evp_y = []
# 初始化 openpyxl
wb = Workbook()
ws = wb.active

# 进行线性相关所用函数
def linefit(x, y):
    N = float(len(x))
    sx, sy, sxx, syy, sxy = 0, 0, 0, 0, 0
    for i in range(0, int(N)):
        sx += x[i]
        sy += y[i]
        sxx += x[i] * x[i]
        syy += y[i] * y[i]
        sxy += x[i] * y[i]
    a = (sy * sx / N - sxy) / (sx * sx / N - sxx)
    b = (sy - a * sx) / N
    r = abs(sy * sx / N - sxy) / math.sqrt((sxx - sx * sx / N) * (syy - sy * sy / N))
    return a, b, r

# 判断数据是否存在，若不存在则为上下数据的均值，若下一行还不存在则为上两行数据均值
def average(l, arr_cell):
    for col in ws.iter_cols(min_col=4, max_col=len(arr_cell)):
        for cell in col:
            if cell.row != 1 and cell.row != 2:
                if cell.value == 32766:
                    if ws.cell(row=cell.row + 1, column=ord(cell.column) - 64).value == 32766 or None:
                        cell.value = int((ws.cell(row=cell.row - 1, column=ord(cell.column) - 64).value + ws.cell(
                            row=cell.row - 2, column=ord(cell.column) - 64).value)) / 2
                    elif ws.cell(row=cell.row - 1, column=ord(cell.column) - 64).value != -1 and ws.cell(
                            row=cell.row - 2, column=ord(cell.column) - 64).value != -1:
                        cell.value = int((ws.cell(row=cell.row - 1, column=ord(cell.column) - 64).value + ws.cell(
                            row=cell.row + 1, column=ord(cell.column) - 64).value)) / 2
                    else:
                        cell.value = -1
            else:
                if cell.value == 32766:
                    cell.value = -1

# 主函数
def saveExcel(station):
    for dirpath, dirnames, filenames in os.walk('./'):
        for k in range(len(filenames)):
            filename = dirpath + '//' + filenames[k]
            for l in range(len(factorArray) - 1):
                if factorArray[l] in filename:
                    factor = factorArray[l]
                    openTxt = open(filename, 'r', encoding='UTF-8')
                    readTxt = openTxt.read()
                    arr_line = readTxt.split('\n') #将数据提取为一行一行
                    for i in range(len(arr_line) - 1):
                        arr_data = arr_line[i].split() # 某行全部数据
                        arr_data = [int(arr_data) for arr_data in arr_data if arr_data] # 将字符串转换为数组
                        if arr_data[0] == station:
                            arr_cell = arr_data[factorDict[factor][1]:factorDict[factor][2]]
                            
                            # 降水需要将8-20、20-8两行数据删除
                            if l == 3:
                                del arr_cell[3]
                                del arr_cell[3] #会退格
                            for j in range(len(arr_cell)): # 进行初步数据处理
                                if l == 0 and (j == 3 or j == 4 or j == 5):
                                    if arr_cell[j] == 32766:
                                        arr_cell[j] = -1
                                    elif arr_cell[j] > 20000:
                                        arr_cell[j] = (arr_cell[j] - 20000) * 0.1
                                    else:
                                        arr_cell[j] *= 0.1
                                if l == 1 and (j == 3 or j == 4 or j == 5):
                                    if arr_cell[j] == 32766:
                                        arr_cell[j] = 32766
                                    else:
                                        arr_cell[j] /= 10
                                if l == 2 and (j == 3 or j == 4):
                                    if arr_cell[j] == 32766:
                                        arr_cell[j] = -1
                                    else:
                                        arr_cell[j] = arr_cell[j] * 0.01
                                if l == 3 and (j == 3):
                                    if arr_cell[j] < 2000:
                                        arr_cell[j] /= 10
                                    elif arr_cell[j] == 32700:
                                        arr_cell[j] = 0
                                    elif arr_cell[j] > 32000 and arr_cell[j] < 33000:
                                        arr_cell[j] = 0
                                    elif arr_cell[j] > 31000 and arr_cell[j] < 32000:
                                        arr_cell[j] = (arr_cell[j] - 31000) / 10
                                    elif arr_cell[j] > 30000 and arr_cell[j] < 31000:
                                        arr_cell[j] = (arr_cell[j] - 30000) / 10
                                if l == 4 and (j == 3):
                                    if arr_cell[j] == 32766 and arr_cell[j + 1] == 32766:
                                        arr_cell[j] = -1
                                    elif arr_cell[j] == 32766 and arr_cell[j + 1] != 32766:
                                        arr_cell[j] = -2
                                    else:
                                        if arr_cell[j] != 32766 and arr_cell[j + 1] != 32766:
                                            arr_evp_y.append(arr_cell[j])
                                            arr_evp_x.append(arr_cell[j + 1])
                                        if arr_cell[j] > 1000:
                                            arr_cell[j] = (arr_cell[j] - 1000) / 10
                                        else:
                                            arr_cell[j] /= 10
                                if l == 5 and (j == 3 or j == 4 or j == 5 or j == 6 or j == 7):
                                    if arr_cell[j] == 32766:
                                        arr_cell[j] = -1
                                if l == 6 and (j == 3):
                                    if arr_cell[j] == 32766:
                                        arr_cell[j] = 32766
                                    else:
                                        arr_cell[j] /= 10
                                ws.cell(row=factorDict[factor][3] + 1, column=j + 1, value=arr_cell[j]) #写入数据
                            factorDict[factor][3] += 1 #一行写完（j）行数加一

                            # 在每个表格完成后进行二次数据处理
                            if arr_cell[0] == endDate[0] and arr_cell[1] == endDate[1] and arr_cell[2] == endDate[2]:
                                if l == 1:
                                    average(l, arr_cell)
                                elif l == 3:
                                    for col in ws.iter_cols(min_col=5):
                                        for cell in col:
                                            cell.value = None
                                elif l == 4: #线性相关
                                    p, q, r = linefit(arr_evp_x, arr_evp_y)
                                    for col in ws.iter_cols(min_col=4, max_col=4):
                                        for cell in col:
                                            if cell.value == -2:
                                                cell.value = int((p * ws.cell(row=cell.row, column=ord(
                                                    cell.column) - 64 + 1).value + q)) / 10
                                    for col in ws.iter_cols(min_col=5): #清空第五行
                                        for cell in col:
                                            cell.value = None
                                elif l == 6:
                                    average(l, arr_cell)
                                name = [''] * (len(factorArray) - 1)
                                name[l] = stationDict[str(station)] + '_' + factorDict[factor][0] + '.xlsx'
                                wb.save(name[l]) #保存文件
                                print("已完成：")
                                print(name[l])
                                for col in ws.iter_cols(min_col=1): #一张表完成后清空整张表格
                                    for cell in col:
                                        cell.value = None
                                factorDict[factor][3] = 0 # 写入行数归零


for a in range(len(stationArray)):
    stationNum = stationArray[a]
    saveExcel(stationNum)
